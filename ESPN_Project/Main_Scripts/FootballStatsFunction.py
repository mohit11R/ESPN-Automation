import  pandas as pd
from openpyxl import Workbook , load_workbook
from openpyxl.chart import BarChart , Reference

def getData():
    url_teamRank = "https://www.espn.in/football/standings/_/league/eng.1/season/2021"

    table = pd.concat([pd.read_html(url_teamRank)[0] , pd.read_html(url_teamRank)[1]])
    print(table)
    table.to_excel("data_football.xlsx")
    return True



wb = load_workbook("data_football.xlsx")
sheet = wb['Sheet1']


min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# print(min_column,max_column,min_row,max_row)

barChart = BarChart()

data = Reference(sheet , min_col = min_column , max_col = max_column +1 ,min_row = min_row , max_row = max_row )

categories = Reference(sheet , min_col = min_column +2 , max_col = min_column +2 ,min_row = min_row +1  , max_row = max_row )

barChart.add_data(data,titles_from_data = True)
barChart.set_categories(categories)

sheet.add_chart(barChart,"B16")

barChart.title = "Football analysis"
barChart.style = 3
wb.save('barchart_1.xlsx')