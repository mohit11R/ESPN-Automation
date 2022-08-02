from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook , load_workbook
from openpyxl.chart import BarChart , Reference
import pandas as pd


# def TeamRanking(driver):
    
#     cricketLink = driver.find_element_by_xpath("/html/body/div[5]/div[2]/header/nav[1]/ul/li[1]/a/span/span[1]")
#     cricketLink.click()
        
#     rankingInfo = driver.find_element_by_xpath("/html/body/div[5]/div[2]/header/nav[2]/div/ul/li[6]/a/span[1]")
#     rankingInfo.click()
    
#     driver.implicitly_wait(10)
    
#     url = "https://www.espncricinfo.com/rankings/content/page/211271.html"
#     result = requests.get(url)
#     soup = BeautifulSoup(result.text , 'html.parser')
    
#     table = soup.find("table", attrs={"class":"StoryengineTable"})
#     headings = [th.get_text() for th in table.find("tr").find_all("th")]
    
#     datasets = []
#     for row in table.find_all("tr")[1:]:
#         dataset = [td.get_text() for td in row.find_all("td")]
#         datasets.append(dataset)

#     print(datasets)
    
    
  

import pandas as pd
  
def storingData():
    
# Storing ICC test Ranking
    url_teamRank = "https://www.espncricinfo.com/rankings/content/page/211271.html"
    
    table = pd.read_html(url_teamRank)[0]
    
    print(table)
    table.to_excel("data_ICC_Test.xlsx")
    


# storingData()

wb = load_workbook("data_ICC_Test.xlsx")
sheet = wb['Sheet1']


min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

barChart = BarChart()

data = Reference(sheet , min_col = min_column+1 , max_col = max_column,min_row = min_row , max_row = max_row )

categories = Reference(sheet , min_col = min_column+1 , max_col = min_column,min_row = min_row +1 , max_row = max_row )

barChart.add_data(data,titles_from_data = True)
barChart.set_categories(categories)

sheet.add_chart(barChart,"B16")

barChart.title = "Team ICC Test Ranking"
barChart.style = 3
wb.save('barchart.xlsx')






